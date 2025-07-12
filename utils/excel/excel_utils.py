# utils/excel_utils.py
# Excel操作相关的工具函数

import logging


class ExcelColumnHelper:
    """Excel列定位助手，支持通过列名定位列位置"""
    
    def __init__(self, ws, header_row=2):
        """
        初始化列定位助手
        Args:
            ws: openpyxl worksheet对象
            header_row: 表头所在行号（默认第2行）
        """
        self.ws = ws
        self.header_row = header_row
        self.columns = {}
        self._build_column_mapping()
    
    def _build_column_mapping(self):
        """构建列名到列索引的映射"""
        try:
            for idx, cell in enumerate(self.ws[self.header_row]):
                if cell.value and str(cell.value).strip():
                    clean_name = str(cell.value).strip()
                    self.columns[clean_name] = idx
            logging.info(f"成功映射 {len(self.columns)} 个Excel列")
        except Exception as e:
            logging.error(f"构建列映射时出错: {e}")
    
    def get_col_index(self, column_name):
        """
        获取列索引，支持容错
        Args:
            column_name: 列名
        Returns:
            int or None: 列索引，如果列不存在返回None
        """
        return self.columns.get(column_name.strip())
    
    def has_column(self, column_name):
        """
        检查列是否存在
        Args:
            column_name: 列名
        Returns:
            bool: 列是否存在
        """
        return column_name.strip() in self.columns
    
    def safe_write(self, row, column_name, value):
        """
        安全写入数据到指定列
        Args:
            row: Excel行对象
            column_name: 列名
            value: 要写入的值
        Returns:
            bool: 是否写入成功
        """
        col_idx = self.get_col_index(column_name)
        if col_idx is not None:
            try:
                row[col_idx].value = value if value is not None else None
                return True
            except Exception as e:
                logging.error(f"写入列 '{column_name}' 时出错: {e}")
                return False
        else:
            logging.warning(f"列 '{column_name}' 不存在，跳过写入")
            return False
    
    def safe_write_hyperlink(self, row_num, column_name, url, display_text=None):
        """
        安全写入超链接到指定列
        Args:
            row_num: 行号
            column_name: 列名
            url: 链接地址
            display_text: 可选的显示文本，如果不提供则使用URL作为显示文本
        Returns:
            bool: 是否写入成功
        """
        col_idx = self.get_col_index(column_name)
        if col_idx is not None and url:
            try:
                # openpyxl中列索引从1开始，但我们的索引从0开始，所以要+1
                cell = self.ws.cell(row=row_num, column=col_idx + 1)
                cell.hyperlink = url
                # 设置显示文本，如果没有提供则使用URL
                cell.value = display_text if display_text else url
                return True
            except Exception as e:
                logging.error(f"写入超链接到列 '{column_name}' 时出错: {e}")
                return False
        else:
            if col_idx is None:
                logging.warning(f"列 '{column_name}' 不存在，跳过超链接写入")
            return False


def safe_write_cell(cell, value):
    """
    通用的安全写入单元格函数
    Args:
        cell: Excel单元格对象
        value: 要写入的值
    """
    try:
        cell.value = value if value is not None else None
    except Exception as e:
        logging.error(f"写入单元格时出错: {e}")


def get_workbook_info(file_path):
    """
    获取工作簿基本信息
    Args:
        file_path: Excel文件路径
    Returns:
        dict: 包含工作簿信息的字典
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active
        return {
            'worksheet_name': ws.title,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'target_year': str(ws['A1'].value)[:4] if ws['A1'].value else None
        }
    except Exception as e:
        logging.error(f"获取工作簿信息时出错: {e}")
        return {} 