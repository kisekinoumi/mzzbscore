# -*- coding: utf-8 -*-
import sys
import os

# 设置UTF-8编码，确保在exe环境中正确处理中文字符
if sys.platform == 'win32':
    import io
    if hasattr(sys.stdout, 'buffer'):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    if hasattr(sys.stderr, 'buffer'):
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 设置环境变量
os.environ['PYTHONIOENCODING'] = 'utf-8'

# 表格模板格式，如果修改值要求使用者更新表格文件
FORMAT_VERSION = 20260410

import time
from html import unescape

import pandas as pd
from openpyxl import load_workbook

# 导入自定义模块
from models import Anime
from utils import preprocess_name, setup_logger, date_error, UrlChecker, setup_twitter_config
from utils.core.global_variables import FILE_PATH, update_constants
from utils.network import setup_proxy, get_proxy_status, is_twitter_accessible, check_update
from src.extractors import (
    extract_bangumi_data,
    extract_myanimelist_data,
    extract_anilist_data,
    extract_filmarks_data
)
from src.data_process.excel_handler import update_excel_data
from utils import ExcelColumnHelper
import concurrent.futures

# 配置日志
logging = setup_logger()

# 第一步：代理检测和配置（程序运行的第一步）
try:
    proxy_config = setup_proxy()
    if proxy_config:
        logging.info(f"✅ 代理配置完成 - {get_proxy_status()}")
    else:
        logging.info(f"📡 网络配置完成 - {get_proxy_status()}")
except Exception as e:
    logging.error(f"代理配置过程中出现错误: {e}")
    logging.info("程序将使用直连模式继续运行")

# 检查更新（仅在exe环境下）
check_update()

# 输出分隔线，明确标识代理配置完成
logging.info("=" * 50)

# 强制清空日期错误列表，确保每次运行都是干净的开始
date_error.clear()

# 防止重复执行的标志
if __name__ != "__main__":
    exit()

wb = None  # 初始化wb变量

try:
    logging.info("程序开始运行...")
    
    # 配置Twitter粉丝数获取功能
    twitter_config_success = False
    
    # 检查Twitter是否可用
    if not is_twitter_accessible():
        logging.warning("Twitter网络不可用，跳过Twitter粉丝数获取功能配置")
        twitter_config_success = False
    else:
        try:
            twitter_config_success = setup_twitter_config()
            if not twitter_config_success:
                logging.warning("Twitter配置失败，将跳过Twitter粉丝数获取功能")
        except Exception as e:
            logging.error(f"Twitter配置过程中出现错误: {e}")
            logging.info("程序将继续运行其他功能")
            twitter_config_success = False
    
    # 输出分隔线，明确标识Twitter配置完成
    logging.info("📋 开始处理动画数据...")
    
    # 读取Excel文件
    try:
        wb = load_workbook(FILE_PATH)
        logging.info(f"成功加载Excel文件: {FILE_PATH}")
    except Exception as e:
        logging.error(f"无法加载Excel文件 {FILE_PATH}: {e}")
        logging.error("请检查Excel文件是否存在且格式正确")
        raise
    ws = wb.active

    # 检查表格格式版本
    excel_version = ws['M1'].value
    if excel_version != FORMAT_VERSION:
        logging.error(f"表格模板版本不匹配！当前代码要求表格文件模板版本为 {FORMAT_VERSION}，但表格模板版本为 {excel_version}。请更新表格模板后重试。")
        logging.error("请访问: https://github.com/kisekinoumi/mzzbscore/releases 下载最新版本的表格模板。")
        raise SystemExit(1)

    # 更新全局常量
    update_constants(str(ws['A1'].value)[:4])  # 读取表格设置目标放送年份

    df = pd.read_excel(FILE_PATH, skiprows=1)
    
    # 清空日期错误列表，避免重复累积
    date_error.clear()
    
    # 创建Excel列助手（只创建一次，避免重复输出映射日志）
    col_helper = ExcelColumnHelper(ws)
    
    # 遍历DataFrame中的每一行数据
    for index, row in df.iterrows():
        if pd.isna(row['原名']):
            logging.warning(f"Skipping row {index} because the original name is NaN.")
            continue  # 跳过这一行
        
        anime = Anime(original_name=row['原名'])  # 获取每行的"原名"列作为原始名称
        logging.info(str(anime))
        
        # 获取当前行的Excel行对象用于链接检查
        excel_row = ws[index + 3]  # DataFrame从0开始，Excel从1开始，且有表头，所以+3
        
        # 检查当前行是否已有链接数据
        existing_urls = UrlChecker.check_row_urls(excel_row, col_helper)
        
        # 如果找到链接，预先设置到anime对象中
        if existing_urls['bangumi']:
            anime.bangumi_url = existing_urls['bangumi']
        if existing_urls['anilist']:
            anime.anilist_url = existing_urls['anilist']
        if existing_urls['myanimelist']:
            anime.myanimelist_url = existing_urls['myanimelist']
        if existing_urls['filmarks']:
            anime.filmarks_url = existing_urls['filmarks']
        
        # 判断是否有任何现有链接
        has_existing_links = UrlChecker.has_any_url(existing_urls)
        available_platforms = UrlChecker.get_available_platforms(existing_urls)
        
        if has_existing_links:
            logging.info(f"发现已有链接的平台: {', '.join(available_platforms)}")
        else:
            logging.info("未发现已有链接，将进行搜索模式")
        
        # 预处理名称（仍然需要，用于没有链接的平台）
        processed_name = preprocess_name(anime.original_name)

        # 提取数据 - 并发执行
        # 原始的串行执行代码（注释掉）
        # extract_bangumi_data(anime, processed_name)
        # extract_myanimelist_data(anime, processed_name)
        # extract_anilist_data(anime, processed_name)
        # extract_filmarks_data(anime, processed_name)
        
        # 创建线程池执行器
        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
            # 提交所有任务到线程池
            future_to_extractor = {
                executor.submit(extract_bangumi_data, anime, processed_name): "bangumi",
                executor.submit(extract_myanimelist_data, anime, processed_name): "myanimelist",
                executor.submit(extract_anilist_data, anime, processed_name): "anilist",
                executor.submit(extract_filmarks_data, anime, processed_name): "filmarks"
            }
            
            # 等待所有任务完成
            for future in concurrent.futures.as_completed(future_to_extractor):
                extractor_name = future_to_extractor[future]
                try:
                    result = future.result()
                    logging.info(f"{extractor_name} extractor completed")
                except Exception as exc:
                    logging.error(f"{extractor_name} extractor generated an exception: {exc}")

        # 如果 MAL 没有找到候选条目，但 AniList 搜到名称，则用 AniList 返回的名称重新搜索 MAL
        mal_not_found = anime.score_mal in ["No acceptable subject found", "No results found"]
        if mal_not_found and anime.anilist_name:
            logging.info("MAL候选未找到，尝试使用 AniList 返回的名称重新搜索 MAL")
            new_processed_name = preprocess_name(anime.anilist_name)
            extract_myanimelist_data(anime, new_processed_name)
        # 如果 AniList 没有找到候选条目，但 MAL 搜到名称，则用 MAL 返回的名称重新搜索 AniList
        anilist_not_found = anime.score_al in ["No acceptable subject found", "No AniList results"]
        if anilist_not_found and anime.myanimelist_name:
            logging.info("AniList候选未找到，尝试使用 MAL 返回的名称重新搜索 AniList")
            new_processed_name = unescape(preprocess_name(anime.myanimelist_name))
            extract_anilist_data(anime, new_processed_name)


        # 获取Twitter粉丝数（如果找到了Twitter账号且配置成功且网络可用）
        if hasattr(anime, 'twitter_username') and anime.twitter_username:
            if not is_twitter_accessible():
                logging.info(f"发现Twitter账号 @{anime.twitter_username}，但Twitter网络不可用，跳过粉丝数获取")
                anime.twitter_followers = "网络不可用"
            elif twitter_config_success:
                try:
                    from src.extractors import TwitterFollowersHelper
                    followers_count = TwitterFollowersHelper.get_followers_count(anime.twitter_username)
                    if followers_count is not None:
                        anime.twitter_followers = followers_count
                    else:
                        logging.warning(f"无法获取 @{anime.twitter_username} 的粉丝数")
                        anime.twitter_followers = "获取失败"
                except Exception as e:
                    logging.error(f"获取Twitter粉丝数时出错: {e}")
                    anime.twitter_followers = "获取出错"
            else:
                logging.info(f"发现Twitter账号 @{anime.twitter_username}，但Twitter配置未成功，跳过粉丝数获取")
                anime.twitter_followers = "配置未成功"

        # 更新Excel数据
        update_excel_data(ws, index, anime, col_helper)

        # 延时以避免频繁请求被拒绝
        time.sleep(0.1)

except Exception as e:
    logging.error(f"发生错误: {e}")

finally:
    # 保存Excel文件
    if wb is not None:
        try:
            wb.save(FILE_PATH)
            logging.info("Excel表格已成功更新。")
        except Exception as e:
            logging.error(f"保存Excel文件时发生错误: {e}")
    else:
        logging.warning("Excel文件未成功加载，跳过保存操作")
    
    # 输出日期错误信息
    try:
        if date_error:
            logging.info("\n" + "=" * 50)
            logging.info("日期错误汇总 (共 %d 条):" % len(date_error))
            logging.info("=" * 50)
            for i, error in enumerate(date_error, 1):
                logging.info("%d. 作品：%s" % (i, error["name"]))
                logging.info("   错误：%s" % error["error"])
                logging.info("-" * 50)
        else:
            logging.info("没有发现任何日期错误！")
    except Exception as e:
        logging.error(f"输出日期错误信息时发生错误: {e}")
    
    # 等待用户输入退出
    try:
        while True:
            user_input = input("输入 'exit' 退出程序: ")
            if user_input.lower() == 'exit':  # 忽略大小写，允许 'exit' 退出
                break
        logging.info("程序已退出...")
    except KeyboardInterrupt:
        logging.info("程序被用户中断...")
    except Exception as e:
        logging.error(f"退出处理时发生错误: {e}")
        logging.info("程序异常退出...")